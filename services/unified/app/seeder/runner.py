"""
seeder/runner.py — Unified startup seeder.

Runs on every app startup (idempotent):
  1. Auth seed: org, 4 demo users, 7 roles, permissions, user-role links.
  2. Warehouse topology seed from warehouse_database.xlsx.

Both are safe to run on a populated DB (ON CONFLICT DO NOTHING / DO UPDATE).
"""
from __future__ import annotations

import os
import re
import uuid
from pathlib import Path
from typing import Optional

import structlog

from app.config import settings
from app.database import AsyncSessionLocal

logger = structlog.get_logger(__name__)

# The xlsx is copied into the Docker image at /app/seed_data/warehouse_database.xlsx
# See Dockerfile COPY instruction.
_XLSX_CANDIDATES = [
    Path("/app/seed_data/warehouse_database.xlsx"),
    Path(__file__).parent.parent / "seed_data" / "warehouse_database.xlsx",
    Path("/warehouse_database.xlsx"),
    # Fallback: look in repo root (useful for local docker-compose.single.yml run)
    Path(__file__).parents[5] / "warehouse_database.xlsx",
]

DEFAULT_WAREHOUSE_ID = settings.WAREHOUSE_ID
DEFAULT_ORG_ID = uuid.UUID("00000000-0000-0000-0000-000000000001")


# ── Auth seed (async, uses SQLAlchemy) ─────────────────────────────────────────

async def seed_auth() -> None:
    """Seed org, roles, permissions, and demo users using the auth-service logic."""
    if not settings.ENABLE_SEEDER:
        return
    try:
        from app.seed import seed_initial_data
        await seed_initial_data()
        logger.info("seeder.auth_complete")
    except Exception as exc:
        logger.error("seeder.auth_failed", error=str(exc))


# ── Warehouse topology seed (sync psycopg2 wrapped in asyncio.to_thread) ───────

def _find_xlsx() -> Optional[Path]:
    for p in _XLSX_CANDIDATES:
        if p.exists():
            return p
    return None


def _seed_warehouse_sync(database_url: str) -> None:
    """
    Synchronous seed of warehouse topology from xlsx using psycopg2.
    Called via asyncio.to_thread so it doesn't block the event loop.
    """
    try:
        import psycopg2
        from openpyxl import load_workbook
    except ImportError as exc:
        logger.warning("seeder.warehouse_deps_missing", error=str(exc))
        return

    xlsx_path = _find_xlsx()
    if xlsx_path is None:
        logger.warning("seeder.xlsx_not_found", searched=[str(p) for p in _XLSX_CANDIDATES])
        return

    # Convert asyncpg URL → plain psycopg2 URL
    db_url = database_url.replace("postgresql+asyncpg://", "postgresql://", 1)

    def make_uuid5(prefix: str, name: str) -> str:
        return str(uuid.uuid5(uuid.UUID("6ba7b810-9dad-11d1-80b4-00c04fd430c8"), f"{prefix}:{name}"))

    def get_field(record: dict, *keys: str, default: str = "") -> str:
        for key in keys:
            if key in record and record[key] is not None:
                v = str(record[key]).strip()
                if v:
                    return v
            nk = key.lower().replace(" ", "").replace("_", "")
            for rk, rv in record.items():
                if rk and rv is not None:
                    nrk = str(rk).lower().replace(" ", "").replace("_", "")
                    if nrk == nk:
                        v = str(rv).strip()
                        if v:
                            return v
        return default

    def extract_loc(code: str) -> dict:
        m = re.match(r"WH-A(\d+)-R(\d+)-RK(\d+)-S(\d+)-P(\d+)", code)
        if not m:
            return {}
        return {"aisle": int(m.group(1)), "row": int(m.group(2)),
                "rack": int(m.group(3)), "shelf": int(m.group(4)),
                "product_pos": int(m.group(5))}

    try:
        wb = load_workbook(str(xlsx_path), read_only=True)
        ws = wb.active
        headers: list[str] = []
        products: list[dict] = []
        for ri, row in enumerate(ws.iter_rows(values_only=True)):
            if ri == 0:
                headers = [str(h).strip() if h else "" for h in row]
                continue
            if not any(v is not None for v in row):
                continue
            products.append({headers[ci]: val for ci, val in enumerate(row) if ci < len(headers)})
        wb.close()
    except Exception as exc:
        logger.error("seeder.xlsx_parse_failed", error=str(exc))
        return

    try:
        conn = psycopg2.connect(db_url)
    except Exception as exc:
        logger.error("seeder.db_connect_failed", error=str(exc))
        return

    try:
        cur = conn.cursor()

        # Warehouse
        cur.execute("""
            INSERT INTO warehouses (id, code, name, address, city, country, total_area_sqm, timezone)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s) ON CONFLICT DO NOTHING
        """, (DEFAULT_WAREHOUSE_ID, "WH-001", "Primary Distribution Center",
              "123 Industrial Blvd", "Bangalore", "India", 25000.00, "Asia/Kolkata"))

        aisles_seen: set = set()
        racks_seen: set = set()
        shelves_seen: set = set()
        bins_to_create: list = []

        for p in products:
            code = get_field(p, "Product_Code", "Product Code", "product_code")
            qr   = get_field(p, "QR_Code", "QR Code", "qr_code")
            loc  = extract_loc(code)
            if not loc:
                continue
            aisles_seen.add(loc["aisle"])
            racks_seen.add((loc["aisle"], loc["rack"]))
            shelves_seen.add((loc["aisle"], loc["rack"], loc["shelf"]))
            bins_to_create.append((loc, qr))

        zone_id = make_uuid5("zone", f"WH-{DEFAULT_WAREHOUSE_ID}-main")
        cur.execute("""
            INSERT INTO zones (id, warehouse_id, code, name, zone_type)
            VALUES (%s,%s,%s,%s,%s) ON CONFLICT (warehouse_id, code) DO NOTHING
        """, (zone_id, DEFAULT_WAREHOUSE_ID, "ZONE-MAIN", "Main Storage Zone", "STORAGE"))

        aisle_ids: dict[int, str] = {}
        for an in sorted(aisles_seen):
            aid = make_uuid5("aisle", f"WH-{DEFAULT_WAREHOUSE_ID}-A{an}")
            cur.execute("""
                INSERT INTO aisles (id, zone_id, code, aisle_number)
                VALUES (%s,%s,%s,%s) ON CONFLICT (zone_id, code) DO NOTHING
            """, (aid, zone_id, f"A{an}", an))
            aisle_ids[an] = aid

        rack_ids: dict[tuple, str] = {}
        for an, rn in sorted(racks_seen):
            rid = make_uuid5("rack", f"WH-{DEFAULT_WAREHOUSE_ID}-A{an}-RK{rn}")
            cur.execute("""
                INSERT INTO racks (id, aisle_id, code, rack_number, num_shelves)
                VALUES (%s,%s,%s,%s,%s) ON CONFLICT (aisle_id, code) DO NOTHING
            """, (rid, aisle_ids[an], f"A{an}-RK{rn}", rn, 4))
            rack_ids[(an, rn)] = rid

        shelf_ids: dict[tuple, str] = {}
        for an, rn, sn in sorted(shelves_seen):
            sid = make_uuid5("shelf", f"WH-{DEFAULT_WAREHOUSE_ID}-A{an}-RK{rn}-S{sn}")
            cur.execute("""
                INSERT INTO shelves (id, rack_id, code, level_number)
                VALUES (%s,%s,%s,%s) ON CONFLICT (rack_id, level_number) DO NOTHING
            """, (sid, rack_ids[(an, rn)], f"A{an}-RK{rn}-S{sn}", sn))
            shelf_ids[(an, rn, sn)] = sid

        bin_ids: dict[str, str] = {}
        for loc, qr in bins_to_create:
            a, rk, s, pos = loc["aisle"], loc["rack"], loc["shelf"], loc["product_pos"]
            bcode = f"A{a}-RK{rk}-S{s}-B{pos}"
            bid = make_uuid5("bin", f"WH-{DEFAULT_WAREHOUSE_ID}-{bcode}")
            sid = shelf_ids.get((a, rk, s))
            if not sid:
                continue
            cur.execute("""
                INSERT INTO bins (id, shelf_id, code, bin_number, qr_code)
                VALUES (%s,%s,%s,%s,%s)
                ON CONFLICT (code) DO UPDATE SET qr_code=EXCLUDED.qr_code, updated_at=NOW()
            """, (bid, sid, bcode, pos, qr or None))
            bin_ids[bcode] = bid

        # Products + inventory
        for p in products:
            code = get_field(p, "Product_Code", "Product Code", "product_code")
            serial = get_field(p, "Product_Serial_Number", "Product Serial Number")
            category = get_field(p, "Category_Number", "Category Number")
            if not code:
                continue
            cur.execute("""
                INSERT INTO products (sku, name, category, unit_of_measure, barcode_value)
                VALUES (%s,%s,%s,%s,%s)
                ON CONFLICT (sku) DO UPDATE SET name=EXCLUDED.name, category=EXCLUDED.category, updated_at=NOW()
            """, (code, f"Product {code}", category or None, "EACH", serial or None))

            loc = extract_loc(code)
            if not loc:
                continue
            a, rk, s, pos = loc["aisle"], loc["rack"], loc["shelf"], loc["product_pos"]
            bcode = f"A{a}-RK{rk}-S{s}-B{pos}"
            bid = bin_ids.get(bcode)
            if bid:
                cur.execute("""
                    INSERT INTO inventory (bin_id, sku, expected_qty)
                    VALUES (%s,%s,%s)
                    ON CONFLICT (bin_id, sku) DO UPDATE SET expected_qty=EXCLUDED.expected_qty, updated_at=NOW()
                """, (bid, code, 1))

        conn.commit()
        cur.close()
        logger.info("seeder.warehouse_complete",
                    aisles=len(aisle_ids), racks=len(rack_ids),
                    shelves=len(shelf_ids), bins=len(bin_ids),
                    products=len(products))
    except Exception as exc:
        conn.rollback()
        logger.error("seeder.warehouse_failed", error=str(exc))
    finally:
        conn.close()


async def seed_warehouse() -> None:
    """Run the synchronous xlsx seeder in a thread pool so it doesn't block the event loop."""
    if not settings.ENABLE_SEEDER:
        return
    import asyncio
    await asyncio.to_thread(_seed_warehouse_sync, settings.DATABASE_URL)
